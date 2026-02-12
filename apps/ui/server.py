from __future__ import annotations

import argparse
import cgi
import csv
import io
import json
import os
from datetime import date, datetime
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any

from apps.agent.main import FinanceAgentConfig, run_finance_agent
from apps.mcp_server.categorization import categorize_merchant
from apps.mcp_server.storage import FinanceStorage
from apps.mcp_server.tools import UploadTransactionsInput, upload_transactions


PROJECT_ROOT = Path(__file__).resolve().parents[2]
STATIC_DIR = PROJECT_ROOT / "apps" / "ui" / "static"
DEFAULT_DB_PATH = PROJECT_ROOT / "data" / "finance.db"

MERCHANT_TRANSLATIONS = {
    "מסטרקרד": "Mastercard",
    "העב' לאחר-נייד": "Post-mobile transfer",
    "העברה-נייד": "Mobile transfer",
    "תשלום שובר": "Voucher payment",
    "משיכה מבנקט": "ATM withdrawal",
    'הו"ק הלואה קרן': "Loan principal debit",
    'הו"ק הלו\' רבית': "Loan interest debit",
    "ג'י וואן פתרונ": "G One Solutions",
    'מופ"ת מילואים': "Reserve duty payment",
    "זיכוי מלאומי": "Leumi credit",
    "בטוח לאומי חד": "National Insurance benefit",
    "bit העברת כסף": "bit money transfer",
    "מקס איט פיננסי": "MAX IT Finance",
    "ריבית מפקדון": "Deposit interest",
    "משיכת פקדון": "Deposit withdrawal",
    "הפקדה לפקדון": "Deposit funding",
    "הפק.שיק בסלולר": "Mobile check deposit",
    "מענק ללקוח": "Customer grant",
}

CATEGORY_LABELS = {
    "grocery": "Grocery",
    "subscriptions": "Subscriptions",
    "transport": "Transport",
    "card_payment": "Card payments",
    "cash_withdrawal": "Cash withdrawal",
    "transfers": "Transfers",
    "loan_interest": "Loan interest",
    "loan_principal": "Loan principal",
    "savings_deposit": "Savings and deposits",
    "benefits_income": "Benefits income",
    "other": "Other",
}

NON_CONSUMPTION_CATEGORIES = {"transfers", "savings_deposit", "loan_principal", "card_payment"}


def _load_env_file(path: Path) -> None:
    if not path.exists():
        return
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def _write_report_file(markdown: str, dataset_id: str, month: str | None) -> str:
    out_dir = PROJECT_ROOT / "output" / "reports"
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_month = month or "all"
    path = out_dir / f"finance_report_{dataset_id[:8]}_{safe_month}_{stamp}.md"
    path.write_text(markdown, encoding="utf-8")
    return str(path)


class FinanceUIHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args: Any, **kwargs: Any) -> None:
        super().__init__(*args, directory=str(STATIC_DIR), **kwargs)

    def do_POST(self) -> None:  # noqa: N802
        if self.path == "/api/run-report":
            self._handle_run_report()
            return

        self._send_json({"error": f"Unknown endpoint: {self.path}"}, status=HTTPStatus.NOT_FOUND)

    def _handle_run_report(self) -> None:
        try:
            payload = self._read_request_payload()
        except ValueError as exc:
            self._send_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            return

        try:
            response = self._run_pipeline(payload)
        except Exception as exc:  # broad to keep UI stable
            self._send_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            return

        self._send_json(response)

    def _read_request_payload(self) -> dict[str, Any]:
        content_type_raw = self.headers.get("Content-Type") or ""
        content_type = content_type_raw.lower()
        length = int(self.headers.get("Content-Length", "0"))
        if length <= 0:
            raise ValueError("Missing request body")

        raw = self.rfile.read(length)

        if content_type.startswith("application/json"):
            try:
                return json.loads(raw.decode("utf-8"))
            except json.JSONDecodeError as exc:
                raise ValueError("Body must be valid JSON") from exc

        if content_type.startswith("multipart/form-data"):
            return _parse_multipart_form_data(content_type=content_type_raw, body=raw)

        raise ValueError("Unsupported content type. Use application/json or multipart/form-data")

    def _run_pipeline(self, payload: dict[str, Any]) -> dict[str, Any]:
        return run_pipeline(payload)

    def _send_json(self, payload: dict[str, Any], status: HTTPStatus = HTTPStatus.OK) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(int(status))
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)


def main() -> int:
    parser = argparse.ArgumentParser(description="Run basic Finance Agent web UI")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=8080)
    parser.add_argument("--env-file", default=".env")
    args = parser.parse_args()

    _load_env_file(PROJECT_ROOT / args.env_file)

    server = ThreadingHTTPServer((args.host, args.port), FinanceUIHandler)
    print(f"Finance UI running at http://{args.host}:{args.port}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()
    return 0


def run_pipeline(payload: dict[str, Any]) -> dict[str, Any]:
    db_path = str(DEFAULT_DB_PATH)
    dataset_id = str(payload.get("dataset_id", "")).strip() or None
    csv_path = str(payload.get("csv_path", "")).strip() or None
    csv_text = str(payload.get("csv_text", "")).strip() or None
    upload_bytes = payload.get("upload_bytes")
    upload_filename = str(payload.get("upload_filename", "")).strip() or None
    requested_month = str(payload.get("month", "")).strip() or None
    recommendations = int(payload.get("recommendations", 3))
    use_llm = _to_bool(payload.get("use_llm", True))
    llm_model = str(payload.get("llm_model", "gpt-4o-mini")).strip() or "gpt-4o-mini"

    if upload_bytes:
        csv_text = _convert_uploaded_to_csv_text(
            filename=upload_filename or "uploaded.csv",
            content=upload_bytes,
        )

    if not dataset_id and not csv_path and not csv_text:
        raise ValueError("Provide dataset_id, csv_path, or file upload")

    upload_result: dict[str, Any] | None = None
    if csv_text:
        uploaded = upload_transactions(
            UploadTransactionsInput(
                csv_text=csv_text,
                source_name=upload_filename or "uploaded_file",
                db_path=db_path,
            )
        )
        dataset_id = uploaded.dataset_id
        upload_result = uploaded.model_dump()
    elif csv_path:
        input_path = Path(csv_path)
        if not input_path.exists():
            raise ValueError(f"CSV path not found: {csv_path}")
        csv_text_from_path = input_path.read_text(encoding="utf-8")
        uploaded = upload_transactions(
            UploadTransactionsInput(
                csv_text=csv_text_from_path,
                source_name=input_path.name,
                db_path=db_path,
            )
        )
        dataset_id = uploaded.dataset_id
        upload_result = uploaded.model_dump()

    if recommendations < 3 or recommendations > 7:
        raise ValueError("recommendations must be between 3 and 7")

    month_context = _resolve_month_context(
        db_path=db_path,
        dataset_id=dataset_id,
        requested_month=requested_month,
    )
    selected_month = month_context["selected_month"]

    config = FinanceAgentConfig(db_path=db_path, use_llm=use_llm, llm_model=llm_model)
    result = run_finance_agent(
        dataset_id=dataset_id,
        month=selected_month,
        recommendations=recommendations,
        config=config,
    )

    report_path = _write_report_file(result["final_markdown"], dataset_id=dataset_id, month=selected_month)

    enriched_top_merchants = []
    for item in result["top_merchants"]["top_merchants"]:
        merchant = str(item.get("merchant", ""))
        category, _ = categorize_merchant(merchant, "")
        enriched_top_merchants.append(
            {
                **item,
                "merchant_en": translate_merchant(merchant),
                "category": category,
            }
        )

    filtered_top_merchants = [
        item for item in enriched_top_merchants if item["category"] not in NON_CONSUMPTION_CATEGORIES
    ]
    if not filtered_top_merchants:
        filtered_top_merchants = enriched_top_merchants

    category_breakdown = []
    for item in result["monthly_report"]["category_breakdown"]:
        category = str(item.get("category", "other"))
        category_breakdown.append(
            {
                **item,
                "category_label": CATEGORY_LABELS.get(category, category.replace("_", " ").title()),
            }
        )

    adjusted_spent = round(
        sum(
            float(item["amount"])
            for item in category_breakdown
            if str(item["category"]) not in NON_CONSUMPTION_CATEGORIES
        ),
        2,
    )
    raw_spent = float(result["monthly_report"]["total_spent"])
    total_income = float(result["monthly_report"]["total_income"])
    adjusted_net = round(total_income - adjusted_spent, 2)
    is_saving = adjusted_net >= 0

    filtered_suggestions = [
        item
        for item in result["budget_suggestions"]["suggestions"]
        if str(item.get("category", "other")) not in NON_CONSUMPTION_CATEGORIES
    ]
    if len(filtered_suggestions) < 3:
        filtered_suggestions = result["budget_suggestions"]["suggestions"]

    return {
        "dataset_id": dataset_id,
        "month": selected_month,
        "month_requested": requested_month,
        "data_range": month_context,
        "upload_result": upload_result,
        "report_path": report_path,
        "monthly_report": {
            **result["monthly_report"],
            "category_breakdown": category_breakdown,
            "total_spent_raw": raw_spent,
            "total_spent_adjusted": adjusted_spent,
            "total_expenses_core": adjusted_spent,
            "total_income_all": total_income,
            "net_balance_adjusted": adjusted_net,
            "savings_or_loss": adjusted_net,
            "is_saving": is_saving,
            "calculation_formula": "savings_or_loss = total_income_all - total_expenses_core",
            "spend_mode": "adjusted_excludes_transfers_deposits_loan_principal_card_payment",
        },
        "top_merchants": {
            **result["top_merchants"],
            "top_merchants": filtered_top_merchants,
        },
        "budget_suggestions": result["budget_suggestions"],
        "budget_suggestions_ui": {
            **result["budget_suggestions"],
            "suggestions": filtered_suggestions,
        },
        "final_markdown": result["final_markdown"],
        "ui_labels": {"categories": CATEGORY_LABELS},
    }


def translate_merchant(name: str) -> str:
    cleaned = name.strip()
    if cleaned in MERCHANT_TRANSLATIONS:
        return MERCHANT_TRANSLATIONS[cleaned]
    return cleaned


def _to_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes", "on"}
    return bool(value)


def _parse_multipart_form_data(*, content_type: str, body: bytes) -> dict[str, Any]:
    payload: dict[str, Any] = {}
    environ = {
        "REQUEST_METHOD": "POST",
        "CONTENT_TYPE": content_type,
        "CONTENT_LENGTH": str(len(body)),
    }
    form = cgi.FieldStorage(
        fp=io.BytesIO(body),
        environ=environ,
        keep_blank_values=True,
    )

    if not form.list:
        return payload

    for field in form.list:
        name = field.name
        if not name:
            continue
        if field.filename:
            payload["upload_filename"] = field.filename
            payload["upload_bytes"] = field.file.read()
        else:
            payload[name] = (field.value or "").strip()

    return payload


def _convert_uploaded_to_csv_text(*, filename: str, content: bytes) -> str:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        return _decode_csv_bytes(content)
    if suffix in {".xlsx", ".xls"}:
        return _excel_bytes_to_csv_text(content=content, suffix=suffix)
    raise ValueError("Uploaded file must be .csv, .xlsx, or .xls")


def _decode_csv_bytes(content: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "cp1255", "iso-8859-8"):
        try:
            return content.decode(enc)
        except UnicodeDecodeError:
            continue
    raise ValueError("Unable to decode CSV file. Please save as UTF-8 CSV.")


def _excel_bytes_to_csv_text(*, content: bytes, suffix: str) -> str:
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise ValueError("openpyxl is required for .xlsx uploads") from exc

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        output = io.StringIO()
        writer = csv.writer(output)
        for row in ws.iter_rows(values_only=True):
            writer.writerow([_cell_to_text(cell) for cell in row])
        wb.close()
        return output.getvalue()

    # .xls fallback via pandas, if available
    try:
        import pandas as pd
    except Exception as exc:
        raise ValueError("pandas is required for .xls uploads") from exc

    try:
        frame = pd.read_excel(io.BytesIO(content))
    except Exception as exc:
        raise ValueError("Failed to parse .xls file") from exc
    return frame.to_csv(index=False)


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _resolve_month_context(*, db_path: str, dataset_id: str, requested_month: str | None) -> dict[str, Any]:
    storage = FinanceStorage(db_path)
    storage.initialize()
    rows = storage.fetch_transactions(dataset_id=dataset_id, month=None)
    if not rows:
        raise ValueError("No transactions found for dataset")

    months = sorted({str(row["txn_date"])[:7] for row in rows})
    first_date = min(str(row["txn_date"]) for row in rows)
    last_date = max(str(row["txn_date"]) for row in rows)
    latest_month = months[-1]

    if requested_month and requested_month not in months:
        raise ValueError(
            f"Requested month {requested_month} is not in dataset range {months[0]}..{months[-1]}"
        )

    selected_month = requested_month or latest_month

    return {
        "selected_month": selected_month,
        "latest_month": latest_month,
        "first_date": first_date,
        "last_date": last_date,
        "months": months,
    }


if __name__ == "__main__":
    raise SystemExit(main())
